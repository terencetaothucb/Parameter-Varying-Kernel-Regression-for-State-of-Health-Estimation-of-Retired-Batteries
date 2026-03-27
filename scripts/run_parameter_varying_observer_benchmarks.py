from __future__ import annotations

from pathlib import Path
from typing import Callable

import numpy as np
from openpyxl import Workbook

from observer_benchmark_utils import (
    AggData,
    GlobalObserver,
    SparseScheduledObserver,
    conditional_linearity_rows,
    load_aggregated_20ah,
    sqdist,
    standardize_apply,
    standardize_fit,
)


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "results" / "benchmark_results_20ah.xlsx"


def metrics(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    err = y_pred - y_true
    mse = np.mean(err**2)
    var = np.mean((y_true - y_true.mean()) ** 2)
    ape = np.abs(err) / np.abs(y_true) * 100.0
    return {
        "mae": float(np.mean(np.abs(err))),
        "median_ape_pct": float(np.median(ape)),
        "r2": float(1.0 - mse / var if var > 0 else 0.0),
        "rmse": float(np.sqrt(mse)),
    }


class ParameterVaryingObserver:
    def __init__(
        self,
        gamma: float,
        lam: float,
        soc_scale: float,
        n_centers: int,
        name: str = "Ours",
    ) -> None:
        self.gamma = gamma
        self.lam = lam
        self.soc_scale = soc_scale
        self.n_centers = n_centers
        self.name = name
        self.mean: np.ndarray | None = None
        self.std: np.ndarray | None = None
        self.centers: np.ndarray | None = None
        self.beta: np.ndarray | None = None

    def _raw(self, data: AggData, mask: np.ndarray) -> np.ndarray:
        return np.column_stack([data.features[mask], self.soc_scale * data.soc[mask] / 100.0])

    def _select_center_idx(self, data: AggData, train_mask: np.ndarray, n_train: int) -> np.ndarray:
        max_centers = min(self.n_centers, n_train)
        train_knots = np.array(sorted(set(data.soc[train_mask])), dtype=float)
        train_idx = np.flatnonzero(train_mask)
        per_knot = max(1, max_centers // max(1, len(train_knots)))
        picked_global: list[int] = []

        for knot in train_knots:
            local_idx = np.flatnonzero(train_mask & np.isclose(data.soc, knot))
            if len(local_idx) <= per_knot:
                picked_global.extend(local_idx.tolist())
            else:
                select = np.linspace(0, len(local_idx) - 1, per_knot, dtype=int)
                picked_global.extend(local_idx[select].tolist())

        picked_global = sorted(set(picked_global))
        pos_map = {idx: i for i, idx in enumerate(train_idx)}
        picked_local = np.array([pos_map[idx] for idx in picked_global if idx in pos_map], dtype=int)
        if len(picked_local) == 0:
            return np.linspace(0, n_train - 1, max_centers, dtype=int)
        if len(picked_local) > max_centers:
            picked_local = picked_local[:max_centers]
        return picked_local

    def fit(self, data: AggData, train_mask: np.ndarray) -> None:
        X = self._raw(data, train_mask)
        self.mean, self.std = standardize_fit(X)
        Xs = standardize_apply(X, self.mean, self.std)
        center_idx = self._select_center_idx(data, train_mask, len(Xs))
        self.centers = Xs[center_idx]
        Phi = np.exp(-self.gamma * sqdist(Xs, self.centers))
        X1 = np.column_stack([np.ones(len(Phi)), Phi])
        reg = self.lam * np.eye(X1.shape[1])
        reg[0, 0] = 0.0
        self.beta = np.linalg.solve(X1.T @ X1 + reg, X1.T @ data.soh[train_mask])

    def predict(self, data: AggData, mask: np.ndarray) -> np.ndarray:
        assert self.mean is not None and self.std is not None and self.centers is not None and self.beta is not None
        X = self._raw(data, mask)
        Xs = standardize_apply(X, self.mean, self.std)
        Phi = np.exp(-self.gamma * sqdist(Xs, self.centers))
        X1 = np.column_stack([np.ones(len(Phi)), Phi])
        return X1 @ self.beta


def benchmark_definitions() -> list[dict[str, str]]:
    return [
        {"model": "B0-Global-H-Linear", "meaning": "Global linear regression on pulse-test features only; ignores SOC entirely."},
        {"model": "B1-Global-HS-Linear", "meaning": "Global linear regression on pulse-test features and SOC, with SOC treated as an ordinary covariate."},
        {"model": "B2-Global-H-Poly", "meaning": "Global polynomial regression on pulse-test features only; tests whether generic nonlinearity can replace SOC-aware modeling."},
        {"model": "B3-Global-HS-Poly", "meaning": "Global polynomial regression on pulse-test features and SOC; the strongest conventional nonlinear baseline without operating-point continuation."},
        {"model": "B4-Discrete-Local", "meaning": "Local observers fitted at seen SOC knots and switched by nearest operating point during testing."},
        {"model": "Ours", "meaning": "Parameter-varying observer with SOC-balanced centers distributed over the discrete SOC operating-point grid."},
    ]


def common_support_mask(data: AggData) -> np.ndarray:
    return (data.source == "fixed") & (data.soc >= 5.0) & (data.soc <= 60.0)


def protocol_spec() -> dict[str, dict[str, object]]:
    return {
        "hold_6_alt": {
            "heldout_soc": [10.0, 15.0, 25.0, 35.0, 45.0, 55.0],
            "observer": {"gamma": 0.01, "lam": 0.01, "soc_scale": 5.0, "n_centers": 400},
        },
        "hold_7_dense": {
            "heldout_soc": [10.0, 20.0, 25.0, 35.0, 40.0, 50.0, 55.0],
            "observer": {"gamma": 0.02, "lam": 0.01, "soc_scale": 3.0, "n_centers": 200},
        },
        "hold_8_dense": {
            "heldout_soc": [10.0, 15.0, 20.0, 25.0, 35.0, 40.0, 50.0, 55.0],
            "observer": {"gamma": 0.005, "lam": 0.001, "soc_scale": 0.25, "n_centers": 150},
        },
    }


def build_masks(data: AggData, heldout_soc: list[float]) -> tuple[np.ndarray, np.ndarray]:
    common = common_support_mask(data)
    heldout_arr = np.array(heldout_soc, dtype=float)
    train_mask = common & ~np.isin(data.soc, heldout_arr)
    test_mask = common & np.isin(data.soc, heldout_arr)
    return train_mask, test_mask


def model_factories(data: AggData, train_mask: np.ndarray, params: dict[str, object]) -> list[Callable[[], object]]:
    train_knots = np.array(sorted(set(data.soc[train_mask])), dtype=float)
    return [
        lambda: GlobalObserver(use_soc=False, poly=False, name="B0-Global-H-Linear"),
        lambda: GlobalObserver(use_soc=True, poly=False, name="B1-Global-HS-Linear"),
        lambda: GlobalObserver(use_soc=False, poly=True, name="B2-Global-H-Poly"),
        lambda: GlobalObserver(use_soc=True, poly=True, name="B3-Global-HS-Poly"),
        lambda: SparseScheduledObserver(train_knots, topk=3, name="B4-Discrete-Local"),
        lambda: ParameterVaryingObserver(name="Ours", **params),
    ]


def run_protocols(data: AggData) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    all_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    param_rows: list[dict[str, object]] = []

    for protocol_name, spec in protocol_spec().items():
        heldout_soc = spec["heldout_soc"]
        observer_params = spec["observer"]
        train_mask, test_mask = build_masks(data, heldout_soc)
        train_knots = sorted(int(x) for x in sorted(set(data.soc[train_mask])))
        test_knots = sorted(int(x) for x in sorted(set(data.soc[test_mask])))
        protocol_rows: list[dict[str, object]] = []

        for build_model in model_factories(data, train_mask, observer_params):
            model = build_model()
            model.fit(data, train_mask)
            pred = model.predict(data, test_mask)
            row = {
                "protocol": protocol_name,
                "train_knots": ",".join(map(str, train_knots)),
                "test_knots": ",".join(map(str, test_knots)),
                "train_count": int(train_mask.sum()),
                "test_count": int(test_mask.sum()),
                "model": model.name,
                **metrics(data.soh[test_mask], pred),
            }
            protocol_rows.append(row)
            all_rows.append(row)

        ranked = sorted(protocol_rows, key=lambda row: (-row["r2"], row["mae"]))
        best_model = ranked[0]
        best_baseline = next(row for row in ranked if row["model"] != "Ours")
        summary_rows.append(
            {
                "protocol": protocol_name,
                "best_model": best_model["model"],
                "best_r2": best_model["r2"],
                "best_mae": best_model["mae"],
                "best_median_ape_pct": best_model["median_ape_pct"],
                "best_baseline": best_baseline["model"],
                "best_baseline_r2": best_baseline["r2"],
                "r2_gain_vs_best_baseline": best_model["r2"] - best_baseline["r2"],
            }
        )
        param_rows.append(
            {
                "protocol": protocol_name,
                "model": "Ours",
                "gamma": observer_params["gamma"],
                "lam": observer_params["lam"],
                "soc_scale": observer_params["soc_scale"],
                "n_centers": observer_params["n_centers"],
            }
        )

    return all_rows, summary_rows, param_rows


def build_main_table(all_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    main_protocol = "hold_8_dense"
    ordered_models = [
        "B0-Global-H-Linear",
        "B1-Global-HS-Linear",
        "B2-Global-H-Poly",
        "B3-Global-HS-Poly",
        "B4-Discrete-Local",
        "Ours",
    ]
    rows = [row for row in all_rows if row["protocol"] == main_protocol]
    by_name = {row["model"]: row for row in rows}
    table = []
    for model_name in ordered_models:
        row = by_name[model_name]
        table.append(
            {
                "protocol": main_protocol,
                "model": model_name,
                "mae": row["mae"],
                "median_ape_pct": row["median_ape_pct"],
                "r2": row["r2"],
            }
        )
    return table


def append_sheet(wb: Workbook, name: str, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(["empty"])
        return
    columns = list(rows[0].keys())
    ws.append(columns)
    for row in rows:
        ws.append([row[column] for column in columns])


def main() -> None:
    data = load_aggregated_20ah()
    all_rows, summary_rows, param_rows = run_protocols(data)
    main_table_rows = build_main_table(all_rows)
    conditional_rows = conditional_linearity_rows(data)

    summary_rows_meta = [
        {"item": "dataset", "value": "20Ah retired LFP batteries"},
        {"item": "processed_data", "value": "data/processed/data4model_20ah.xlsx"},
        {"item": "shared_support_region", "value": "SOC 5% to 60%"},
        {"item": "main_protocol", "value": "hold_8_dense"},
        {"item": "reported_metrics", "value": "MAE, Median APE (%), R2"},
    ]

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    append_sheet(wb, "Summary", summary_rows_meta)
    append_sheet(wb, "BenchmarkDefs", benchmark_definitions())
    append_sheet(wb, "MainResultTable", main_table_rows)
    append_sheet(wb, "HardSOCInterpolation", all_rows)
    append_sheet(wb, "ProtocolSummary", summary_rows)
    append_sheet(wb, "Hyperparameters", param_rows)
    append_sheet(wb, "ConditionalObservability", conditional_rows)
    wb.save(OUTPUT_PATH)

    print(f"Saved {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
