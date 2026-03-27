from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_DATA_DIR = REPO_ROOT / "data" / "raw" / "20Ah_LFP"
OUTPUT_PATH = REPO_ROOT / "data" / "processed" / "data4model_20ah.xlsx"

MODEL_COLUMNS = [
    "capacity_ah",
    "pulse_width_ms",
    "sample_index_within_capacity",
    "sample_group_id",
    "soc_source",
    "source_file",
    "Qn",
    "Q",
    "SOH",
    "SOC",
    "Hyst_M3_0.5C",
    "Hyst_M3_1C",
    "Hyst_M3_1.5C",
    "Hyst_M3_2C",
    "Hyst_M3_2.5C",
]


def parse_pulse_width_ms(file_name: str) -> int:
    match = re.search(r"_W_(\d+)\.xlsx$", file_name)
    if not match:
        raise ValueError(f"Cannot parse pulse width from file name: {file_name}")
    return int(match.group(1))


def sheet_to_records(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws_all = wb["SOC ALL"]
    header = list(next(ws_all.iter_rows(min_row=1, max_row=1, values_only=True)))

    random_rows: set[tuple[object, ...]] = set()
    ws_random = wb["SOC TEST RANDOM"]
    next(ws_random.iter_rows(min_row=1, max_row=1, values_only=True))
    for row in ws_random.iter_rows(min_row=2, values_only=True):
        random_rows.add(tuple(row))

    records: list[dict[str, object]] = []
    next(ws_all.iter_rows(min_row=1, max_row=1, values_only=True))
    for row in ws_all.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, row))
        record["soc_source"] = "random" if tuple(row) in random_rows else "fixed"
        records.append(record)
    return records


def build_records() -> list[dict[str, object]]:
    files = sorted(RAW_DATA_DIR.glob("*.xlsx"))
    if not files:
        raise ValueError(f"No Excel files found in {RAW_DATA_DIR}")

    combined_records: list[dict[str, object]] = []
    reference_keys: list[tuple[object, object, object, object]] | None = None
    for path in files:
        pulse_width_ms = parse_pulse_width_ms(path.name)
        records = sheet_to_records(path)
        base_keys = [(r["Qn"], r["Q"], r["SOH"], r["SOC"]) for r in records]
        if reference_keys is None:
            reference_keys = base_keys
        elif base_keys != reference_keys:
            raise ValueError(f"Sample alignment mismatch in {path.name}")

        for idx, record in enumerate(records, start=1):
            combined_records.append(
                {
                    "capacity_ah": 20,
                    "pulse_width_ms": pulse_width_ms,
                    "sample_index_within_capacity": idx,
                    "sample_group_id": f"20Ah_{idx:04d}",
                    "soc_source": record["soc_source"],
                    "source_file": path.name,
                    "Qn": record["Qn"],
                    "Q": record["Q"],
                    "SOH": record["SOH"],
                    "SOC": record["SOC"],
                    "Hyst_M3_0.5C": record["Hyst_M3_0.5C"],
                    "Hyst_M3_1C": record["Hyst_M3_1C"],
                    "Hyst_M3_1.5C": record["Hyst_M3_1.5C"],
                    "Hyst_M3_2C": record["Hyst_M3_2C"],
                    "Hyst_M3_2.5C": record["Hyst_M3_2.5C"],
                }
            )

    combined_records.sort(
        key=lambda row: (
            int(row["sample_index_within_capacity"]),
            int(row["pulse_width_ms"]),
        )
    )
    return combined_records


def main() -> None:
    records = build_records()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data4Model"
    ws.append(MODEL_COLUMNS)
    for record in records:
        ws.append([record[column] for column in MODEL_COLUMNS])

    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(records)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
