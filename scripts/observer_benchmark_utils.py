from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = REPO_ROOT / "data" / "processed" / "data4model_20ah.xlsx"

PULSE_ORDER = [30, 50, 70, 100, 300, 500, 700, 1000, 3000, 5000]
HYST_COLS = ["Hyst_M3_0.5C", "Hyst_M3_1C", "Hyst_M3_1.5C", "Hyst_M3_2C", "Hyst_M3_2.5C"]


@dataclass
class AggData:
    sample_group: np.ndarray
    soc: np.ndarray
    source: np.ndarray
    soh: np.ndarray
    features: np.ndarray
    feature_labels: list[str]
    fixed_knots: np.ndarray


def load_aggregated_20ah() -> AggData:
    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    ws = wb["Data4Model"]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {name: i for i, name in enumerate(header)}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    points: dict[tuple[str, float, str], dict[str, object]] = {}
    feature_labels: list[str] = []
    for row in rows:
        key = (row[idx["sample_group_id"]], float(row[idx["SOC"]]), row[idx["soc_source"]])
        point = points.setdefault(
            key,
            {
                "sample_group": row[idx["sample_group_id"]],
                "SOC": float(row[idx["SOC"]]),
                "source": row[idx["soc_source"]],
                "SOH": float(row[idx["SOH"]]),
                "pulse_to_hyst": {},
            },
        )
        pulse = int(row[idx["pulse_width_ms"]])
        point["pulse_to_hyst"][pulse] = [float(row[idx[name]]) for name in HYST_COLS]

    aggregated = []
    for point in points.values():
        pulse_to_hyst = point["pulse_to_hyst"]
        if sorted(pulse_to_hyst) != PULSE_ORDER:
            continue
        features = []
        labels = []
        for pulse in PULSE_ORDER:
            vals = pulse_to_hyst[pulse]
            features.extend(vals)
            labels.extend([f"{pulse}_{name}" for name in HYST_COLS])
        feature_labels = labels
        aggregated.append(
            (
                point["sample_group"],
                point["SOC"],
                point["source"],
                point["SOH"],
                np.array(features, dtype=float),
            )
        )

    sample_group = np.array([row[0] for row in aggregated])
    soc = np.array([row[1] for row in aggregated], dtype=float)
    source = np.array([row[2] for row in aggregated])
    soh = np.array([row[3] for row in aggregated], dtype=float)
    features = np.row_stack([row[4] for row in aggregated])
    fixed_knots = np.array(sorted(set(soc[source == "fixed"])), dtype=float)

    return AggData(
        sample_group=sample_group,
        soc=soc,
        source=source,
        soh=soh,
        features=features,
        feature_labels=feature_labels,
        fixed_knots=fixed_knots,
    )


def standardize_fit(X: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    mean = X.mean(axis=0)
    std = X.std(axis=0)
    std[std == 0] = 1.0
    return mean, std


def standardize_apply(X: np.ndarray, mean: np.ndarray, std: np.ndarray) -> np.ndarray:
    return (X - mean) / std


def add_poly2(X: np.ndarray) -> np.ndarray:
    return np.column_stack([X, X**2])


def fit_ridge(X: np.ndarray, y: np.ndarray, lam: float = 1e-4) -> np.ndarray:
    X1 = np.column_stack([np.ones(len(X)), X])
    reg = lam * np.eye(X1.shape[1])
    reg[0, 0] = 0.0
    return np.linalg.solve(X1.T @ X1 + reg, X1.T @ y)


def predict_ridge(beta: np.ndarray, X: np.ndarray) -> np.ndarray:
    X1 = np.column_stack([np.ones(len(X)), X])
    return X1 @ beta


def sqdist(A: np.ndarray, B: np.ndarray) -> np.ndarray:
    aa = np.sum(A * A, axis=1, keepdims=True)
    bb = np.sum(B * B, axis=1)[None, :]
    return aa + bb - 2 * A @ B.T


class GlobalObserver:
    def __init__(self, use_soc: bool, poly: bool, name: str) -> None:
        self.use_soc = use_soc
        self.poly = poly
        self.name = name
        self.mean: np.ndarray | None = None
        self.std: np.ndarray | None = None
        self.beta: np.ndarray | None = None

    def _design(self, data: AggData, mask: np.ndarray) -> np.ndarray:
        X = data.features[mask]
        if self.use_soc:
            X = np.column_stack([X, data.soc[mask]])
        if self.poly:
            X = add_poly2(X)
        return X

    def fit(self, data: AggData, train_mask: np.ndarray) -> None:
        X = self._design(data, train_mask)
        self.mean, self.std = standardize_fit(X)
        Xs = standardize_apply(X, self.mean, self.std)
        self.beta = fit_ridge(Xs, data.soh[train_mask], 1e-4)

    def predict(self, data: AggData, mask: np.ndarray) -> np.ndarray:
        assert self.mean is not None and self.std is not None and self.beta is not None
        X = self._design(data, mask)
        Xs = standardize_apply(X, self.mean, self.std)
        return predict_ridge(self.beta, Xs)


class SparseScheduledObserver:
    def __init__(self, knots: np.ndarray, topk: int, name: str) -> None:
        self.knots = knots
        self.topk = topk
        self.name = name
        self.local_models: dict[float, tuple[list[int], np.ndarray]] = {}

    def fit(self, data: AggData, train_mask: np.ndarray) -> None:
        for knot in self.knots:
            local_mask = train_mask & np.isclose(data.soc, knot)
            correlations = []
            for j in range(data.features.shape[1]):
                corr = np.corrcoef(data.features[local_mask, j], data.soh[local_mask])[0, 1]
                correlations.append(0.0 if np.isnan(corr) else abs(corr))
            best_idx = list(np.argsort(-np.array(correlations))[: self.topk])
            X = data.features[local_mask][:, best_idx]
            beta = fit_ridge(X, data.soh[local_mask], 1e-8)
            self.local_models[float(knot)] = (best_idx, beta)

    def _local_pred(self, knot: float, z: np.ndarray) -> float:
        idx, beta = self.local_models[float(knot)]
        return float(predict_ridge(beta, z[idx][None, :])[0])

    def predict(self, data: AggData, mask: np.ndarray) -> np.ndarray:
        preds = []
        for soc_value, z in zip(data.soc[mask], data.features[mask]):
            if soc_value <= self.knots[0]:
                chosen = self.knots[0]
            elif soc_value >= self.knots[-1]:
                chosen = self.knots[-1]
            else:
                right_idx = np.searchsorted(self.knots, soc_value)
                left = self.knots[right_idx - 1]
                right = self.knots[right_idx]
                chosen = left if abs(soc_value - left) <= abs(soc_value - right) else right
            preds.append(self._local_pred(chosen, z))
        return np.array(preds)


def conditional_linearity_rows(data: AggData) -> list[dict[str, object]]:
    rows = []
    fixed_mask = data.source == "fixed"
    for knot in data.fixed_knots:
        mask = fixed_mask & np.isclose(data.soc, knot)
        correlations = []
        for j in range(data.features.shape[1]):
            corr = np.corrcoef(data.features[mask, j], data.soh[mask])[0, 1]
            correlations.append(0.0 if np.isnan(corr) else corr)
        corr_arr = np.array(correlations)
        best_idx = int(np.argmax(np.abs(corr_arr)))
        rows.append(
            {
                "soc_knot": float(knot),
                "count": int(mask.sum()),
                "best_feature": data.feature_labels[best_idx],
                "best_corr": float(corr_arr[best_idx]),
                "best_abs_corr": float(abs(corr_arr[best_idx])),
                "median_abs_corr": float(np.median(np.abs(corr_arr))),
            }
        )
    return rows
